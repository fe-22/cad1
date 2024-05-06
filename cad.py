import tkinter as tk
from tkinter import ttk, messagebox
from PIL import Image, ImageTk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import openpyxl
import os

class GeradorExcel:
    def __init__(self, arquivo_nome, diretorio=None):
        self.arquivo_nome = arquivo_nome
        self.diretorio = diretorio

    def escrever_dados(self, dados):
        caminho_completo = os.path.join(self.diretorio, self.arquivo_nome)
        if not os.path.exists(caminho_completo):
            workbook = openpyxl.Workbook()
            workbook.save(caminho_completo)
        workbook = openpyxl.load_workbook(caminho_completo)
        sheet = workbook.active
        for linha in dados:
            sheet.append(linha)
        print("Dados escritos na planilha:")
        for row in sheet.iter_rows(values_only=True):
            print(row)
        workbook.save(caminho_completo)

    def salvar(self):
        if self.diretorio:
            caminho_completo = os.path.join(self.diretorio, self.arquivo_nome)
            caminho_completo_com_extensao = caminho_completo + ".xlsx"

            if os.path.exists(caminho_completo_com_extensao):
                workbook = openpyxl.load_workbook(caminho_completo_com_extensao)
                for sheet in workbook.sheetnames:
                    workbook[sheet].delete_rows(1, workbook[sheet].max_row)
                workbook.save(caminho_completo_com_extensao)
            else:
                workbook = openpyxl.Workbook()
                workbook.save(caminho_completo_com_extensao)
        else:
            raise ValueError("O diretório de salvamento não foi especificado.")

    def definir_diretorio(self, diretorio):
        self.diretorio = diretorio

def cadastrar_novo_membro():
    try:
        # Obter os valores dos campos
        nome = nome_var.get()
        cpf = cpf_var.get()
        data_nascimento = data_nascimento_var.get()
        sexo = sexo_var.get()
        estado_civil = estado_civil_var.get()
        profissao = profissao_var.get()
        escolaridade = escolaridade_var.get()
        data_batismo = data_batismo_var.get()
        data_batismo_data = data_batismo_data_var.get()
        data_conversao = data_conversao_var.get()
        data_ingresso = data_ingresso_var.get()
        endereco_residencial = endereco_residencial_var.get()
        telefone = telefone_var.get()
        email = email_var.get()
        # Campos adicionais
        cargo_na_igreja = cargo_na_igreja_var.get()
        area_ministerio = area_ministerio_var.get()
        nome_conjuge = nome_conjuge_var.get()
        filhos = filhos_var.get()
        alergias = alergias_var.get()
        doencas_cronicas = doencas_cronicas_var.get()
        observacoes = observacoes_var.get()

        # Criar uma instância da classe GeradorExcel
        gerador_excel = GeradorExcel("example.xlsx", "c:/teste/cad1")

        # Salvar os dados no Excel
        gerador_excel.escrever_dados([[nome, cpf, data_nascimento, sexo, estado_civil, profissao, escolaridade,
                                       data_batismo,data_batismo_data, data_conversao, data_ingresso, endereco_residencial,
                                       telefone, email, cargo_na_igreja, area_ministerio, nome_conjuge,
                                       filhos, alergias, doencas_cronicas, observacoes]])

        # Salvar o arquivo Excel
        gerador_excel.salvar()

        # Atualizar o contador de membros
        atualizar_contador_membros()

        messagebox.showinfo("Sucesso", "Novo membro cadastrado com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao cadastrar o novo membro: {e}")

def atualizar_contador_membros():
    global total_membros
    total_membros += 1
    label_total_membros["text"] = f"Total de Membros: {total_membros}"

class DashboardWindow(tk.Toplevel):
    def __init__(self, master=None):
        super().__init__(master)
        self.title('Comparativos - ADFidelidade')

        # Frame para os campos do formulário
        form_frame = tk.Frame(self)
        form_frame.pack(padx=20, pady=20, fill=tk.BOTH, expand=True)

        # Canvas para exibir o frame dos gráficos com barra de rolagem
        canvas = tk.Canvas(self)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Adicionar uma barra de rolagem ao canvas
        scrollbar = ttk.Scrollbar(self, orient=tk.VERTICAL, command=canvas.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.configure(yscrollcommand=scrollbar.set)

        # Frame para exibir os gráficos
        self.graph_frame = tk.Frame(canvas)
        canvas.create_window((0, 0), window=self.graph_frame, anchor=tk.NW)

        # Configurar a barra de rolagem para o canvas
        self.graph_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    def obter_dados_do_usuario(self):
        # Aqui você define o tamanho do widget do gráfico
        largura_grafico = 400
        altura_grafico = 300

        # Coletar dados dos campos de entrada na interface gráfica
        sexo = sexo_var.get()
        estado_civil = estado_civil_var.get()
        escolaridade = escolaridade_var.get()
        batismo_aguas = data_batismo_var.get()  # Use data_batismo_var em vez de entry_batismo_aguas
        cargo_na_igreja = cargo_na_igreja_var.get()

        # Organizar os dados em uma lista de tuplas
        dados = [
            ("Sexo", sexo),
            ("Estado Civil", estado_civil),
            ("Escolaridade", escolaridade),
            ("Batismo nas Águas", batismo_aguas),
            ("Cargo na Igreja", cargo_na_igreja)
        ]
        return dados, largura_grafico, altura_grafico

    def plotar_graficos(self, dados, largura_grafico, altura_grafico):
        # Limpar o frame do gráfico
        for widget in self.graph_frame.winfo_children():
            widget.destroy()

        # Extrair categorias e valores dos dados
        categorias = [item[0] for item in dados]
        valores = [item[1] for item in dados]

        # Converter valores para float (se possível)
        valores_numericos = []
        for val in valores: 
            try:
                val = float(val)
            except (ValueError, TypeError):
                val = 0  # Atribuir 0 se não for possível converter para float
            valores_numericos.append(val)

        # Plotar um gráfico de pizza para cada categoria
        for i, categoria in enumerate(categorias):
            fig, ax = plt.subplots()
            ax.pie([valores_numericos[i]], labels=[categoria], autopct='%1.1f%%', startangle=90)
            ax.set_title(f'Gráfico {categoria}')

            # Adicionar o gráfico ao frame
            canvas = FigureCanvasTkAgg(fig, master=self.graph_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=10, 
                                         ipadx=largura_grafico, ipady=altura_grafico)

def exibir_dashboard():
    # Criar a janela do dashboard
    dashboard_window = DashboardWindow(app)
    
    # Obter os dados para os gráficos
    dados, largura_grafico, altura_grafico = dashboard_window.obter_dados_do_usuario()

    # Plotar os gráficos no dashboard
    dashboard_window.plotar_graficos(dados, largura_grafico, altura_grafico)


app = tk.Tk()
app.title('ADFidelidade')

# Obtenha a largura e a altura da tela disponível
largura_tela = app.winfo_screenwidth()
altura_tela = app.winfo_screenheight()

# Defina a geometria da janela para preencher a tela
app.geometry(f"{largura_tela}x{altura_tela}")

# Carregar o logotipo
imagem_caminho = r"C:\teste\cad1\logo.png.png"
imagem = Image.open(imagem_caminho)
largura_maxima = 80  # pixels
altura_maxima = 90   # pixels
imagem.thumbnail((largura_maxima, altura_maxima))
imagem_tk = ImageTk.PhotoImage(imagem)

# Cabeçalho 1
header1 = tk.Label(app, text="Cadastro de Membro", font=("Arial", 12, "bold"))
header1.grid(row=0, column=0, columnspan=2, pady=10, sticky="nsew")

# Logo
label_imagem = tk.Label(app, image=imagem_tk)
label_imagem.grid(row=0, column=0, sticky="n")

# Cabeçalho 2
header2 = tk.Label(app, text="ASSEMBLEIA DE DEUS - MINISTÉRIO FIDELIDADE/SBC", font=("Arial", 14, "bold"))
header2.grid(row=1, column=0, columnspan=2, pady=10, sticky="nsew")

# Variáveis para armazenar os valores dos campos
nome_var = tk.StringVar()
cpf_var = tk.StringVar()
data_nascimento_var = tk.StringVar()
sexo_var = tk.StringVar()
estado_civil_var = tk.StringVar()
profissao_var = tk.StringVar()
escolaridade_var = tk.StringVar()
data_batismo_var = tk.StringVar()
data_batismo_data_var = tk.StringVar()
data_conversao_var = tk.StringVar()
data_ingresso_var = tk.StringVar()
endereco_residencial_var = tk.StringVar()
telefone_var = tk.StringVar()
email_var = tk.StringVar()
total_membros = 0

# Campos adicionais
cargo_na_igreja_var = tk.StringVar()
area_ministerio_var = tk.StringVar()
nome_conjuge_var = tk.StringVar()
filhos_var = tk.StringVar()
alergias_var = tk.StringVar()
doencas_cronicas_var = tk.StringVar()
observacoes_var = tk.StringVar()

# Frame para os campos do formulário
form_frame = tk.Frame(app)
form_frame.grid(row=2, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")

# Campos Dados Básicos
label_nome = tk.Label(form_frame, text="Nome Completo:")
label_nome.grid(row=1, column=0, sticky="e", padx=(0, 10))
entry_nome = tk.Entry(form_frame, textvariable=nome_var, width=100)
entry_nome.grid(row=1, column=1)

label_cpf = tk.Label(form_frame, text="CPF:")
label_cpf.grid(row=2, column=0, sticky="e", padx=(10, 10))
entry_cpf = tk.Entry(form_frame, textvariable=cpf_var, width=80)
entry_cpf.grid(row=2, column=1)

label_data_nascimento = tk.Label(form_frame, text="Data de Nascimento:")
label_data_nascimento.grid(row=3, column=0, sticky="e", padx=(0,  10))
entry_data_nascimento = tk.Entry(form_frame, textvariable=data_nascimento_var, width=50)
entry_data_nascimento.grid(row=3, column=1)

label_endereco_residencial = tk.Label(form_frame, text="Endereço Residencial:")
label_endereco_residencial.grid(row=4, column=0, sticky="e", padx=(0, 10))
entry_endereco_residencial = tk.Entry(form_frame, textvariable=endereco_residencial_var, width=100)
entry_endereco_residencial.grid(row=4, column=1)

label_telefone = tk.Label(form_frame, text="Telefone:")
label_telefone.grid(row=5, column=0, sticky="e", padx=(100, 10))
entry_telefone = tk.Entry(form_frame, textvariable=telefone_var, width=50)
entry_telefone.grid(row=5, column=1)

label_email = tk.Label(form_frame, text="Email:")
label_email.grid(row=6, column=0, sticky="e", padx=(0, 10))
entry_email = tk.Entry(form_frame, textvariable=email_var, width=100)
entry_email.grid(row=6, column=1)

# Variável para armazenar o sexo selecionado
sexo_var = tk.StringVar()
sexo_var.set("Masculino")  # Valor padrão inicial

label_sexo = tk.Label(form_frame, text="Sexo:")
label_sexo.grid(row=7, column=0, sticky="e", padx=(0, 10))

# Botões de rádio
radio_masculino = tk.Radiobutton(form_frame, text="Masculino", variable=sexo_var, value="Masculino")
radio_masculino.grid(row=7, column=1, sticky="w")

radio_feminino = tk.Radiobutton(form_frame, text="Feminino", variable=sexo_var, value="Feminino")
radio_feminino.grid(row=8, column=1, sticky="w")

# Combobox para o estado civil
label_estado_civil = tk.Label(form_frame, text="Estado Civil:")
label_estado_civil.grid(row=9, column=0, sticky="nsew", padx=(0, 10))

# Opções para o combobox
opcoes_estado_civil = ["Casado", "Solteiro", "Viúvo", "Divorciado"]

# Combobox
combo_estado_civil = ttk.Combobox(form_frame, values=opcoes_estado_civil, textvariable=estado_civil_var, width=47)
combo_estado_civil.current(0)  # Definir o valor padrão inicial
combo_estado_civil.grid(row=9, column=1, sticky="w")

label_conjuge = tk.Label(form_frame, text="Nome do cônjuge:")
label_conjuge.grid(row=10, column=0, sticky="e", padx=(0, 10))
entry_conjuge = tk.Entry(form_frame, textvariable=nome_conjuge_var, width=100)
entry_conjuge.grid(row=10, column=1)

label_profissao = tk.Label(form_frame, text="Profissão:")
label_profissao.grid(row=11, column=0, sticky="e", padx=(0, 10))
entry_profissao = tk.Entry(form_frame, textvariable=profissao_var, width=100)
entry_profissao.grid(row=11, column=1)

label_escolaridade = tk.Label(form_frame, text="Escolaridade:")
label_escolaridade.grid(row=12, column=0, sticky="e", padx=(0, 10))
entry_escolaridade = tk.Entry(form_frame, textvariable=escolaridade_var, width=100)
entry_escolaridade.grid(row=12, column=1)

header_ministeriais = tk.Label(form_frame, text="Dados Ministeriais ", font=("Arial", 12, "bold"))
header_ministeriais.grid(row=13, columnspan=2, pady=(20, 5))

# Campos Dados Adicionais
label_cargo_na_igreja = tk.Label(form_frame, text="Exerce algum Cargo na Igreja/Qual?:")
label_cargo_na_igreja.grid(row=14, column=0, sticky="e", padx=(0, 10))
entry_cargo_na_igreja = tk.Entry(form_frame, textvariable=cargo_na_igreja_var, width=100)
entry_cargo_na_igreja.grid(row=14, column=1)

# Labels e botões de rádio para o batismo nas águas
label_data_batismo = tk.Label(form_frame, text="Batizado nas águas?:")
label_data_batismo.grid(row=15, column=0, sticky="e", padx=(0, 10))

radio_data_batismo = tk.Radiobutton(form_frame, text="Sim", variable=data_batismo_var, value="Sim")
radio_data_batismo.grid(row=15, column=1, sticky="w")

radio_data_batismo = tk.Radiobutton(form_frame, text="Não", variable=data_batismo_var, value="Não")
radio_data_batismo.grid(row=16, column=1, sticky="w")

label_data_batismo_data = tk.Label(form_frame, text="Data do batismo nas águas?:")
label_data_batismo_data.grid(row=17, column=0, sticky="e", padx=(0, 10))
entry_data_batismo_data = tk.Entry(form_frame, textvariable=data_batismo_data_var, width=100)
entry_data_batismo_data.grid(row=17, column=1)

# Labels e botões de rádio para o batismo no Espírito Santo
label_data_conversao = tk.Label(form_frame, text="Batizado no Espírito Santo?:")
label_data_conversao.grid(row=18, column=0, sticky="e", padx=(0, 10))

data_conversao = tk.Radiobutton(form_frame, text="Sim", variable=data_conversao_var, value="Sim")
data_conversao.grid(row=18, column=1, sticky="w")

data_conversao = tk.Radiobutton(form_frame, text="Não", variable=data_conversao_var, value="Não")
data_conversao.grid(row=19, column=1, sticky="w")

label_data_conversao_data = tk.Label(form_frame, text="Data do batismo no Espírito Santo?:")
label_data_conversao_data.grid(row=20, column=0, sticky="e", padx=(0, 10))

entry_data_conversao_data = tk.Entry(form_frame, textvariable=data_conversao, width=100)
entry_data_conversao_data.grid(row=20, column=1)

label_data_ingresso = tk.Label(form_frame, text="Data de quando chegou em nossa igreja:")
label_data_ingresso.grid(row=21, column=0, sticky="w", padx=(10, 10))
entry_data_ingresso = tk.Entry(form_frame, textvariable=data_ingresso_var, width=50)
entry_data_ingresso.grid(row=21, column=1)

label_observacoes = tk.Label(form_frame, text="Observações:")
label_observacoes.grid(row=22, column=0, sticky="w", padx=(10, 10))
entry_observacoes = tk.Entry(form_frame, textvariable=observacoes_var, width=100)
entry_observacoes.grid(row=22, column=1)

# Label para exibir o total de membros
label_total_membros = tk.Label(app, text="Total de Membros: 0")
label_total_membros.grid(row=0, column=3, padx=10, pady=10, sticky="e")

# Botão para cadastrar um novo membro
button_cadastrar = ttk.Button(app, text="Cadastrar Membro", command=cadastrar_novo_membro)
button_cadastrar.grid(row=17, column=3, padx=10, pady=10, sticky="e")

# Botão para abrir o dashboard
button_dashboard = ttk.Button(app, text="Abrir Dashboard", command=lambda: exibir_dashboard())
button_dashboard.grid(row=17, column=2, padx=10, pady=10, sticky="w")

# Função principal
if __name__ == "__main__":
    app.mainloop()
